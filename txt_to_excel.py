#!/usr/bin/env python3
import os
import glob
import re
import json
import posixpath
import sqlite3
import shutil
import asyncio
import aiohttp
import requests
from datetime import datetime
from urllib.parse import urlparse, parse_qsl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def escape_formula(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')): return "'" + value
    return value

def make_absolute(url, domain):
    if url.startswith('http://') or url.startswith('https://'): return url
    elif url.startswith('//'): return f"https:{url}"
    elif url.startswith('/'): return f"https://{domain}{url}"
    else: return f"https://{domain}/{url}"

def get_status_color(status):
    status_str = str(status)
    if status_str.startswith('2'): return '28A745'
    if status_str.startswith('3'): return '17A2B8'
    if status_str.startswith('4'): return 'FD7E14'
    if status_str.startswith('5'): return 'DC3545'
    if 'Static' in status_str: return 'A8B8D0'
    if 'Skipped' in status_str: return 'E83E8C' 
    if 'Legacy' in status_str: return '6C757D'
    return '6C757D'

def get_status_priority(status):
    s = str(status)
    if s.startswith('2'): return 1
    if s.startswith('5'): return 2
    if s in ['401', '403']: return 3
    if s.startswith('3'): return 4
    if s.startswith('4'): return 5
    return 6

def get_safe_domain(target):
    return "wild_" + target[2:] if target.startswith('*.') else target

def normalize_dynamic_path(path):
    p = re.sub(r'[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}', '{UUID}', path)
    p = re.sub(r'\b\d{3,}\b', '{ID}', p)
    p = re.sub(r'\b[a-zA-Z0-9]{10,}\b', '{HASH}', p)
    return p

# 패시브 정찰 단계용 주요 보안 필터
regex_sensitive_exts = re.compile(r'\.(env|bak|swp|old|sql|sqlite|db|dump|log|config|properties|yml|yaml|ini)$', re.IGNORECASE)
regex_sensitive_paths = re.compile(r'/(admin|administrator|wp-admin|manage|phpmyadmin|server-status|server-info|actuator|swagger-ui|graphql)($|/)', re.IGNORECASE)
regex_credential_params = re.compile(r'(?:\?|&)(api_?key|token|jwt|auth|secret|password|pwd|access_?token)=([a-zA-Z0-9\-_\.]{8,})', re.IGNORECASE)
regex_infra_paths = re.compile(r'/\.(git|svn|hg|aws|ssh|docker)($|/)', re.IGNORECASE)

def get_best_gemini_model(api_key):
    print("[*] 현재 API 키로 사용 가능한 최적의 Gemini AI 모델을 동적으로 탐색합니다...", flush=True)
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
        res = requests.get(url, timeout=15)
        if res.status_code == 200:
            models = res.json().get('models', [])
            valid_models = [m['name'] for m in models if 'generateContent' in m.get('supportedGenerationMethods', [])]
            
            priorities = ['models/gemini-1.5-flash-latest', 'models/gemini-2.5-flash', 'models/gemini-1.5-flash', 'models/gemini-pro']
            for p in priorities:
                if p in valid_models:
                    selected = p.replace('models/', '')
                    print(f"[+] API 승인 확인! 타격 모델 설정 완료: {selected}", flush=True)
                    return selected
            if valid_models: return valid_models[0].replace('models/', '')
    except Exception as e:
        print(f"[-] 모델 동적 탐색 실패: {e}")
    return "gemini-1.5-flash"

async def ask_gemini_async(session, gemini_key, batch, model_name):
    prompt = (
        "You are an elite Bug Bounty Hunter and Red Teamer. Analyze the following list of URLs discovered during passive reconnaissance.\n"
        "Evaluate the probability (0 to 100) that each URL contains a sensitive information disclosure or critical vulnerability based purely on its paths, parameters, and naming conventions.\n"
        "Return EXACTLY a JSON array of objects. Do not include markdown formatting or backticks. Each object must contain these keys:\n"
        "- 'url': the exact URL string\n"
        "- 'probability': integer from 0 to 100\n"
        "- 'vuln_type': string of suspected vulnerability type\n"
        "- 'reason': short clear explanation in Korean of why this URL is high risk.\n\n"
        f"URLs:\n{json.dumps(batch)}"
    )
    g_api_url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={gemini_key}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json"}
    }
    timeout = aiohttp.ClientTimeout(total=150)
    for attempt in range(3):
        try:
            async with session.post(g_api_url, json=payload, timeout=timeout) as res:
                if res.status == 200:
                    res_json = await res.json()
                    raw_reply = res_json.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')
                    if not raw_reply: return []
                    match = re.search(r'\[\s*\{.*?\}\s*\]', raw_reply, re.DOTALL)
                    if match:
                        try: return json.loads(match.group(0))
                        except: pass
                    try: return json.loads(raw_reply.strip())
                    except: return []
                elif res.status == 429: await asyncio.sleep(10); continue
                else: return []
        except: await asyncio.sleep(3); continue
    return []

async def process_all_gemini(gemini_key, candidate_urls, model_name):
    ai_ranked_results = []
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(candidate_urls), 10):
            batch = candidate_urls[i:i+10]
            print(f"[*] Gemini AI 지능형 분석 진행 중... ({i+1} ~ {min(i+10, len(candidate_urls))} / {len(candidate_urls)})")
            res_list = await ask_gemini_async(session, gemini_key, batch, model_name)
            if isinstance(res_list, list): ai_ranked_results.extend(res_list)
            if i + 10 < len(candidate_urls): await asyncio.sleep(3)
    return ai_ranked_results

# ✨ Wayback Machine 연혁 확인 함수
async def fetch_wayback_first_seen(session, subdomain):
    url = f"https://web.archive.org/cdx/search/cdx?url={subdomain}/*&limit=1&fl=timestamp&output=json"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) PassiveRecon/1.0"}
    for attempt in range(3):
        try:
            async with session.get(url, headers=headers, timeout=10) as response:
                if response.status == 200:
                    data = await response.json()
                    if len(data) > 1 and len(data[1]) > 0:
                        ts = data[1][0]
                        return f"{ts[:4]}-{ts[4:6]}-{ts[6:8]}"
                    else:
                        return "기록 없음"
                elif response.status in [429, 503]:
                    await asyncio.sleep(2)
                    continue
                else:
                    break
        except Exception:
            await asyncio.sleep(1)
    return "기록 없음"

# ✨ 실시간 응답(httpx 유사 기능) 확인 함수
async def fetch_subdomain_status(session, subdomain):
    try:
        # HTTPS 우선 시도 (SSL 인증서 오류 무시)
        async with session.get(f"https://{subdomain}", timeout=5, allow_redirects=False) as resp:
            return str(resp.status)
    except Exception:
        try:
            # 실패 시 HTTP 재시도
            async with session.get(f"http://{subdomain}", timeout=5, allow_redirects=False) as resp:
                return str(resp.status)
        except Exception:
            return "Dead"

# ✨ 단일 서브도메인 동시 분석 래퍼 함수
async def analyze_subdomain(session, subdomain):
    wb_task = asyncio.create_task(fetch_wayback_first_seen(session, subdomain))
    status_task = asyncio.create_task(fetch_subdomain_status(session, subdomain))
    
    wb_date = await wb_task
    status = await status_task
    return subdomain, wb_date, status

# ✨ 전체 서브도메인 분석 매니저
async def analyze_all_subdomains(subdomains):
    results = {}
    # SSL 검증을 비활성화하여 빠른 IP 연결 유무만 파악
    connector = aiohttp.TCPConnector(ssl=False, limit=50)
    async with aiohttp.ClientSession(connector=connector) as session:
        tasks = []
        for sub in subdomains:
            tasks.append(analyze_subdomain(session, sub))
            # Rate Limit 및 과부하 방지를 위해 20개씩 끊어서 병렬 처리
            if len(tasks) >= 20:
                batch_results = await asyncio.gather(*tasks)
                for sub, wb_date, status in batch_results:
                    results[sub] = {"wayback": wb_date, "status": status}
                tasks = []
                await asyncio.sleep(1) 
        if tasks:
            batch_results = await asyncio.gather(*tasks)
            for sub, wb_date, status in batch_results:
                results[sub] = {"wayback": wb_date, "status": status}
    return results

def build_advanced_excel_report():
    print("[+] 초고속 SQLite DB 기반 차분 분석(Differential Analysis) 엔진 가동 중...", flush=True)
    
    targets = []
    if os.path.exists('targets.txt'):
        with open('targets.txt', 'r') as f:
            targets = [line.strip() for line in f if line.strip() and not line.strip().startswith('#')]
    
    os.makedirs('reports', exist_ok=True)
    db_path = 'reports/recon_history.db'
    prev_db_path = 'previous_report/recon_history.db'
    
    if os.path.exists(prev_db_path):
        shutil.copy(prev_db_path, db_path)
        
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS master_urls (url TEXT PRIMARY KEY)")
    cursor.execute("CREATE TABLE IF NOT EXISTS downloaded_js (url TEXT PRIMARY KEY)")
    cursor.execute("CREATE TABLE IF NOT EXISTS historical_subdomains (subdomain TEXT PRIMARY KEY)")
    cursor.execute('''CREATE TABLE IF NOT EXISTS target_stats (
        target TEXT PRIMARY KEY,
        passive_tot INTEGER DEFAULT 0,
        jsluice_tot INTEGER DEFAULT 0,
        katana_tot INTEGER DEFAULT 0 
    )''')
    conn.commit()

    cursor.execute("SELECT target FROM target_stats")
    db_targets = [row[0] for row in cursor.fetchall()]
    all_targets = list(set(targets + db_targets))

    target_map = {get_safe_domain(t): t for t in all_targets}
    matrix_data = {raw_target: {} for raw_target in all_targets}
    signature_counts = {}
    
    junk_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.svg', '.css', '.woff', '.woff2', '.ico', '.eot', '.ttf', '.mp4')
    blacklist_words = ['logout', 'signout', 'delete', 'remove', 'revoke', 'destroy']
    
    js_url_converter = {}
    for mf in glob.glob('results/*_js_mapping.txt'):
        try:
            with open(mf, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    if '\t' in line:
                        s, o = line.strip().split('\t', 1)
                        js_url_converter[s] = o
        except: pass

    all_today_raw_urls = []
    temp_file_records = []

    for file_path in glob.glob('results/*.*'):
        filename = os.path.basename(file_path).lower()
        match = re.match(r'^(.*)_(linkfinder|trufflehog|gau|waybackurls|katana)\.txt$', filename)
        if not match: continue
        
        safe_domain = match.group(1)
        if safe_domain not in target_map: continue
        
        raw_target = target_map[safe_domain]
        is_wildcard = raw_target.startswith('*.')
        base_domain = raw_target[2:] if is_wildcard else raw_target

        if 'linkfinder' in filename or 'jsluice' in filename: source_tool = 'LinkFinder'
        elif 'trufflehog' in filename: source_tool = 'TruffleHog'
        elif 'waybackurls' in filename: source_tool = 'Waybackurls'
        elif 'gau' in filename: source_tool = 'GAU'
        else: continue

        try:
            with open(file_path, 'r', errors='ignore') as f:
                for line in f:
                    line_str = line.strip()
                    if not line_str or line_str.startswith('#'): continue
                    line_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', line_str)
                    if not line_str: continue
                    
                    if '\t' in line_str: js_file, raw_url = line_str.split('\t', 1)
                    else: js_file, raw_url = "Passive Archive", line_str
                    
                    if js_file in js_url_converter: js_file = js_url_converter[js_file]
                    
                    abs_url = make_absolute(raw_url, base_domain)
                    parsed_netloc = urlparse(abs_url).netloc.split(':')[0]
                    
                    if is_wildcard:
                        if not (parsed_netloc == base_domain or parsed_netloc.endswith('.' + base_domain)): continue
                    else:
                        if parsed_netloc != base_domain: continue

                    if urlparse(abs_url).path.lower().endswith(junk_extensions): continue
                    
                    all_today_raw_urls.append(abs_url)
                    temp_file_records.append((raw_target, abs_url, source_tool, js_file))
        except: pass

    existing_urls_cache = set()
    all_today_raw_urls = list(set(all_today_raw_urls))
    
    for idx in range(0, len(all_today_raw_urls), 500):
        chunk = all_today_raw_urls[idx:idx+500]
        placeholders = ",".join(["?"] * len(chunk))
        cursor.execute(f"SELECT url FROM master_urls WHERE url IN ({placeholders})", chunk)
        for row in cursor.fetchall(): existing_urls_cache.add(row[0])

    for raw_target, abs_url, source_tool, js_file in temp_file_records:
        parsed_for_sig = urlparse(abs_url)
        query_keys = tuple(sorted([k for k, v in parse_qsl(parsed_for_sig.query, keep_blank_values=True)]))
        norm_path = normalize_dynamic_path(parsed_for_sig.path)
        signature = (parsed_for_sig.netloc, posixpath.dirname(norm_path), posixpath.splitext(norm_path)[1], query_keys)

        if abs_url not in matrix_data[raw_target]:
            if signature_counts.get(signature, 0) >= 5: continue
            signature_counts[signature] = signature_counts.get(signature, 0) + 1
            
            is_new = abs_url not in existing_urls_cache
            matrix_data[raw_target][abs_url] = {"tools": set(), "files": set(), "is_new": is_new}
            
        matrix_data[raw_target][abs_url]["tools"].add(source_tool)
        if source_tool in ['LinkFinder', 'TruffleHog']:
            matrix_data[raw_target][abs_url]["files"].add(js_file)

    status_codes = {}
    for res_file in glob.glob('results/httpx_results_*.json') + glob.glob('strike_out/httpx_results_*.json'):
        try:
            with open(res_file, 'r', errors='ignore') as f:
                for line in f:
                    if not line.strip(): continue
                    data = json.loads(line.strip())
                    status_codes[data.get('url')] = data.get('status_code', 'Dead')
        except: pass

    gemini_key = os.environ.get('GEMINI_API_KEY')
    ai_ranked_results = []
    
    if gemini_key:
        candidate_urls = []
        for url_map in matrix_data.values():
            for url, data in url_map.items():
                sc = str(status_codes.get(url, 'Dead'))
                if 'TruffleHog' in data["tools"] or sc in ['200', '301', '302', '401', '403', '500'] or '?' in url:
                    candidate_urls.append(url)
                    
        candidate_urls = list(set(candidate_urls))[:300]
        if candidate_urls:
            selected_model = get_best_gemini_model(gemini_key)
            print(f"[+] 총 {len(candidate_urls)}개의 중요 엔드포인트를 식별하여 AI 추론을 요청합니다...", flush=True)
            ai_ranked_results = asyncio.run(process_all_gemini(gemini_key, candidate_urls, selected_model))
            
            if ai_ranked_results:
                ai_ranked_results.sort(key=lambda x: x.get('probability', 0), reverse=True)

    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    
    postman_collection = {
        "info": {
            "name": f"🎯 Passive Recon API Collection ({now_str})",
            "description": "과거 아카이브를 통해 추출한 도메인별 API 및 엔드포인트 명세서입니다.",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "item": []
    }

    wb = Workbook()
    font_header, fill_header = Font(name='Malgun Gothic', bold=True, color='FFFFFF'), PatternFill(start_color='2F3542', end_color='2F3542', fill_type='solid')
    font_data, fill_zebra = Font(name='Malgun Gothic', size=10, color='333333'), PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    align_center, align_left = Alignment(horizontal='center', vertical='center'), Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"), top=Side(style="thin", color="E0E0E0"), bottom=Side(style="thin", color="E0E0E0"))

    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    
    dash_headers = [
        "No", "타겟 도메인", "🌟 서브도메인 (누적/신규)", "📊 누적 / 🔥 신규 URL", 
        "jsluice (누적 / 신규)", "TruffleHog 탐지",
        "🟢 200 (OK)", "🟠 403/401 (권한)", "🔴 500대 (에러)"
    ]
    ws_dash.append(dash_headers)
    for c in range(1, len(dash_headers) + 1): 
        ws_dash.cell(1, c).font = font_header; ws_dash.cell(1, c).fill = fill_header
        ws_dash.cell(1, c).alignment = align_center; ws_dash.cell(1, c).border = thin_border

    dash_idx = 2
    all_today_discovered_urls = []
    high_risk_records = []
    
    g_passive_tot = g_passive_new = 0
    g_jsluice_tot = g_jsluice_new = 0
    g_truf = g_200 = g_40x = g_50x = 0

    cursor.execute("SELECT subdomain FROM historical_subdomains")
    previous_subdomains = {row[0] for row in cursor.fetchall()}
    
    global_new_subdomains = set() 
    global_current_subdomains = set() 

    for raw_target, url_map in matrix_data.items():
        cursor.execute("SELECT passive_tot FROM target_stats WHERE target = ?", (raw_target,))
        has_db = cursor.fetchone()
        if not url_map and not has_db: continue

        sheet_title = re.sub(r'[\\/\?\*\:\[\]]', '_', raw_target)[:30]
        postman_folder = {"name": raw_target, "item": []}
        
        today_passive_count = len(url_map)
        domain_new_count = sum(1 for data in url_map.values() if data.get("is_new", False))
        
        today_jsluice_total = sum(1 for data in url_map.values() if 'LinkFinder' in data["tools"])
        jsluice_new = sum(1 for data in url_map.values() if 'LinkFinder' in data["tools"] and data.get("is_new", False))
        
        trufflehog_count = sum(1 for data in url_map.values() if 'TruffleHog' in data["tools"])

        cursor.execute("SELECT passive_tot, jsluice_tot, katana_tot FROM target_stats WHERE target = ?", (raw_target,))
        row = cursor.fetchone()
        db_katana_tot = 0
        if row:
            db_passive_tot, db_jsluice_tot, db_katana_tot = row
            new_passive_tot = db_passive_tot + domain_new_count
            new_jsluice_tot = db_jsluice_tot + jsluice_new
        else:
            new_passive_tot = today_passive_count
            new_jsluice_tot = today_jsluice_total
            
        cursor.execute("INSERT OR REPLACE INTO target_stats (target, passive_tot, jsluice_tot, katana_tot) VALUES (?, ?, ?, ?)", 
                       (raw_target, new_passive_tot, new_jsluice_tot, db_katana_tot))

        count_200 = count_40x = count_50x = 0
        for url in url_map.keys():
            all_today_discovered_urls.append(url)
            status = str(status_codes.get(url, 'Dead'))
            if status.startswith('2'): count_200 += 1
            elif status in ['401', '403']: count_40x += 1
            elif status.startswith('5'): count_50x += 1

        current_subdomains = {urlparse(u).netloc for u in url_map.keys() if urlparse(u).netloc}
        new_subdomains = current_subdomains - previous_subdomains
        
        global_current_subdomains.update(current_subdomains)

        if today_passive_count > 0 and bool(previous_subdomains):
            global_new_subdomains.update(new_subdomains)

        total_sub_count = len(current_subdomains)
        new_sub_count = len(new_subdomains) if bool(previous_subdomains) else 0

        if today_passive_count > 0:
            sub_dash_mark = f"{total_sub_count} / {new_sub_count}"
        else:
            sub_dash_mark = "0 / 0"
        
        g_passive_tot += new_passive_tot; g_passive_new += domain_new_count
        g_jsluice_tot += new_jsluice_tot; g_jsluice_new += jsluice_new
        g_truf += trufflehog_count
        g_200 += count_200; g_40x += count_40x; g_50x += count_50x
        
        ws_dash.append([
            dash_idx - 1, 
            escape_formula(raw_target), 
            sub_dash_mark, 
            f"{new_passive_tot} / {domain_new_count}", 
            f"{new_jsluice_tot} / {jsluice_new}", 
            trufflehog_count, 
            count_200, count_40x, count_50x
        ])
        
        for c in range(1, len(dash_headers) + 1):
            cell = ws_dash.cell(dash_idx, c)
            cell.font = font_data; cell.border = thin_border
            if c == 2:
                if today_passive_count > 0:
                    cell.hyperlink = f"#'{sheet_title}'!A1"
                    cell.font = Font(name='Malgun Gothic', color='0056B3', underline='single')
                else: cell.font = Font(name='Malgun Gothic', color='777777', italic=True)
            elif c == 3 and new_sub_count > 0: 
                cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
            
            if today_passive_count > 0:
                if c == 4 and domain_new_count > 0: cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
                elif c == 5 and jsluice_new > 0: cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
                elif c == 6 and trufflehog_count > 0: cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
            else:
                if c != 2: cell.font = Font(name='Malgun Gothic', color='999999', italic=True)
        dash_idx += 1

        if today_passive_count == 0: continue

        ws = wb.create_sheet(title=sheet_title)
        ws.append(["🔙 대시보드로 돌아가기 (Return to Dashboard)"])
        ws.merge_cells('A1:G1')
        back_cell = ws.cell(row=1, column=1)
        back_cell.hyperlink = "#'Summary Dashboard'!A1"; back_cell.font = Font(name='Malgun Gothic', size=11, bold=True, color='0056B3', underline='single')
        back_cell.fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid'); back_cell.alignment = align_left

        ws.append(["No", "🔥 신규여부", "🌟 신규 서브", "소스 출처", "발견된 JS 파일명", "응답 상태", "타겟 절대 경로 (URL)"])
        for c in range(1, 8): ws.cell(2, c).font = font_header; ws.cell(2, c).fill = fill_header; ws.cell(2, c).alignment = align_center; ws.cell(2, c).border = thin_border

        sorted_urls = sorted(url_map.items(), key=lambda x: (
            not x[1].get("is_new", False), 
            get_status_priority(status_codes.get(x[0], 'Dead') if any(b not in x[0].lower() for b in blacklist_words) else 'Skipped(위험)'), 
            x[0]
        ))
        
        for sub_idx, (url, data) in enumerate(sorted_urls, 1):
            if sub_idx > 1048500: break
            tools_str, files_str = ", ".join(sorted(list(data["tools"]))), ", ".join(sorted(list(data["files"]))) if data["files"] else "-"
            is_new_mark = "🆕 NEW" if data.get("is_new", False) else "-"
            is_blacklist = any(b in url.lower() for b in blacklist_words)
            
            if not is_blacklist:
                parsed_pm = urlparse(url)
                postman_folder["item"].append({"name": parsed_pm.path if parsed_pm.path else "/", "request": {"method": "GET", "header": [], "url": {"raw": url, "protocol": parsed_pm.scheme, "host": parsed_pm.netloc.split('.'), "path": [p for p in parsed_pm.path.split('/') if p], "query": [{"key": k, "value": v} for k, v in parse_qsl(parsed_pm.query, keep_blank_values=True)]}}})

            current_status = "Skipped(위험)" if is_blacklist else ( "Static(생략)" if urlparse(url).path.lower().endswith(junk_extensions) else status_codes.get(url, 'Dead') )
            is_new_subdomain = (urlparse(url).netloc in new_subdomains) and bool(previous_subdomains)
            sub_mark = "🌟 신규" if is_new_subdomain else "-"

            ws.append([sub_idx, is_new_mark, sub_mark, escape_formula(tools_str), escape_formula(files_str), current_status, escape_formula(url)])
            for c in range(1, 8):
                cell = ws.cell(sub_idx + 2, c)
                cell.font = font_data; cell.border = thin_border
                if ((sub_idx+2) % 2) == 1: cell.fill = fill_zebra
                if c == 2 and data.get("is_new", False): cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
                if c == 3 and is_new_subdomain: cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
                if c == 6: cell.fill = PatternFill(start_color=get_status_color(current_status), end_color=get_status_color(current_status), fill_type='solid'); cell.font = Font(name='Malgun Gothic', bold=True, color='FFFFFF'); cell.alignment = align_center
                elif c in [4, 5, 7]: cell.alignment = align_left
                else: cell.alignment = align_center

            is_high_risk, reason = False, ""
            if 'TruffleHog' in data["tools"]: is_high_risk, reason = True, "🔥 [Critical] TruffleHog: 기밀 키(Secret) 유출 의심"
            elif is_blacklist: is_high_risk, reason = True, "⚠️ [Warning] 파괴적 엔드포인트 수동 검점 요망"
            else:
                path_lower = urlparse(url).path.lower()
                if regex_infra_paths.search(path_lower): is_high_risk, reason = True, "🚨 [Infra] 인프라/버전관리 폴더 노출 의심"
                elif regex_sensitive_exts.search(path_lower): is_high_risk, reason = True, "🚨 [File] 민감한 파일 확장자 노출 (백업/설정)"
                elif regex_sensitive_paths.search(path_lower): is_high_risk, reason = True, "🚨 [Path] 관리자/디버그 콘솔 접근 의심"
                elif regex_credential_params.search(urlparse(url).query): is_high_risk, reason = True, "🚨 [Param] 파라미터 내 평문 인증 토큰 포착"

            if is_high_risk:
                high_risk_records.append({
                    "is_new_mark": is_new_mark,
                    "sub_mark": sub_mark,
                    "tools_str": tools_str,
                    "files_str": files_str,
                    "current_status": current_status,
                    "raw_target": raw_target,
                    "url": url,
                    "reason": reason,
                    "is_new": data.get("is_new", False),
                    "is_new_sub": is_new_subdomain,
                    "priority": get_status_priority(current_status)
                })

        if postman_folder["item"]: postman_collection["item"].append(postman_folder)

    # ✨ Wayback Machine 연혁 및 실시간 응답(httpx 유사) 비동기 동시 조회
    subdomain_analysis_results = {}
    if global_current_subdomains:
        print(f"[*] 서브도메인 실시간 응답 상태 및 Wayback 연혁 분석 가동 (총 {len(global_current_subdomains)}개)...", flush=True)
        subdomain_analysis_results = asyncio.run(analyze_all_subdomains(list(global_current_subdomains)))
    
    # ✨ 서브도메인 전용 시트 생성 및 데이터 기록
    if global_current_subdomains:
        ws_subs = wb.create_sheet(title="🌐 서브도메인 연혁(Wayback)")
        ws_subs.append(["No", "서브도메인 (Subdomain)", "🔥 신규 여부", "📡 응답 상태", "최초 발견일 (Wayback Machine)"])
        for c in range(1, 6):
            ws_subs.cell(1, c).font = font_header
            ws_subs.cell(1, c).fill = fill_header
            ws_subs.cell(1, c).alignment = align_center
            ws_subs.cell(1, c).border = thin_border
        
        sub_idx = 2
        # 신규 서브도메인이 상단에 오도록 정렬 후, 알파벳 순 정렬
        sorted_subs = sorted(list(global_current_subdomains), key=lambda x: (x not in global_new_subdomains, x))
        for sub in sorted_subs:
            is_new_mark = "🌟 신규" if sub in global_new_subdomains else "-"
            analysis = subdomain_analysis_results.get(sub, {"wayback": "기록 없음", "status": "Dead"})
            first_seen = analysis["wayback"]
            status = analysis["status"]
            
            ws_subs.append([sub_idx - 1, escape_formula(sub), is_new_mark, status, first_seen])
            for c in range(1, 6):
                cell = ws_subs.cell(sub_idx, c)
                cell.font = font_data; cell.border = thin_border
                if (sub_idx % 2) == 1: cell.fill = fill_zebra
                
                if c == 2: 
                    cell.alignment = align_left
                else: 
                    cell.alignment = align_center
                    
                if c == 3 and sub in global_new_subdomains: 
                    cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
                elif c == 4: 
                    cell.fill = PatternFill(start_color=get_status_color(status), end_color=get_status_color(status), fill_type='solid')
                    cell.font = Font(name='Malgun Gothic', bold=True, color='FFFFFF')
                elif c == 5 and first_seen != "기록 없음": 
                    cell.font = Font(name='Malgun Gothic', bold=True, color='28A745')
            sub_idx += 1

    high_risk_records.sort(key=lambda x: (not x["is_new"], x["priority"], x["raw_target"], x["url"]))
    
    ws_high = wb.create_sheet(title="🚨 High Risk (고위험군)")
    ws_high.append(["No", "🔥 신규여부", "🌟 신규 서브", "소스 출처", "발견된 JS 파일명", "응답 상태", "타겟 도메인", "고위험 경로 (Endpoint)", "탐지 사유"])
    for c in range(1, 10): 
        ws_high.cell(1, c).font = font_header
        ws_high.cell(1, c).fill = fill_header
        ws_high.cell(1, c).alignment = align_center
        ws_high.cell(1, c).border = thin_border

    high_risk_idx = 2
    for hr in high_risk_records:
        ws_high.append([high_risk_idx - 1, hr["is_new_mark"], hr["sub_mark"], escape_formula(hr["tools_str"]), escape_formula(hr["files_str"]), hr["current_status"], escape_formula(hr["raw_target"]), escape_formula(hr["url"]), escape_formula(hr["reason"])])
        for c in range(1, 10):
            cell = ws_high.cell(high_risk_idx, c)
            cell.font = font_data; cell.border = thin_border
            if (high_risk_idx % 2) == 0: cell.fill = fill_zebra
            if c == 2 and hr["is_new"]: cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
            if c == 3 and hr["is_new_sub"]: cell.font = Font(name='Malgun Gothic', bold=True, color='E83E8C')
            if c == 6: cell.fill = PatternFill(start_color=get_status_color(hr["current_status"]), end_color=get_status_color(hr["current_status"]), fill_type='solid'); cell.font = Font(name='Malgun Gothic', bold=True, color='FFFFFF'); cell.alignment = align_center
            elif c in [4, 5, 7, 8, 9]: cell.alignment = align_left
            else: cell.alignment = align_center
        high_risk_idx += 1

    if ai_ranked_results:
        ws_ai = wb.create_sheet(title="🔮 Gemini AI Ranking")
        ws_ai.append(["No", "타겟 URL", "🔮 잠재적 위험 확률", "위험 유형", "Gemini AI 지능형 정보 노출 분석 가이드"])
        for c in range(1, 6):
            ws_ai.cell(1, c).font = font_header
            ws_ai.cell(1, c).fill = PatternFill(start_color='6F42C1', end_color='6F42C1', fill_type='solid')
            ws_ai.cell(1, c).alignment = align_center
            ws_ai.cell(1, c).border = thin_border
        
        for idx, res in enumerate(ai_ranked_results, 1):
            ws_ai.append([idx, escape_formula(res.get('url')), f"{res.get('probability')}%", escape_formula(res.get('vuln_type')), escape_formula(res.get('reason'))])
            for c in range(1, 6):
                cell = ws_ai.cell(idx + 1, c)
                cell.font = font_data; cell.border = thin_border
                if (idx % 2) == 1: cell.fill = fill_zebra
                if c in [2, 4, 5]: cell.alignment = align_left
                else: cell.alignment = align_center

    if all_today_discovered_urls:
        cursor.executemany("INSERT OR IGNORE INTO master_urls (url) VALUES (?)", [(u,) for u in list(set(all_today_discovered_urls))])
        today_subs = {urlparse(u).netloc.split(':')[0] for u in all_today_discovered_urls if urlparse(u).netloc}
        if today_subs:
            cursor.executemany("INSERT OR IGNORE INTO historical_subdomains (subdomain) VALUES (?)", [(s,) for s in today_subs])
        conn.commit()
    conn.close()
    
    if global_new_subdomains:
        with open('reports/new_subdomains.txt', 'w', encoding='utf-8') as f:
            for sub in sorted(list(global_new_subdomains)):
                f.write(sub + '\n')

    if dash_idx > 2:
        g_sub_tot = len(global_current_subdomains)
        g_sub_new = len(global_new_subdomains)
        total_sub_mark = f"{g_sub_tot} / {g_sub_new}"
        
        ws_dash.append([
            "", "📊 총 합계 (Total)", total_sub_mark, 
            f"{g_passive_tot} / {g_passive_new}", 
            f"{g_jsluice_tot} / {g_jsluice_new}", 
            g_truf, g_200, g_40x, g_50x
        ])
        for c in range(1, len(dash_headers) + 1):
            cell = ws_dash.cell(dash_idx, c)
            cell.font = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF'); cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.border = thin_border; cell.alignment = align_center if c != 2 else align_left

    for sheet in wb.worksheets:
        header_row = 1 if sheet.title in ["Summary Dashboard", "🚨 High Risk (고위험군)", "🔮 Gemini AI Ranking", "🌐 서브도메인 연혁(Wayback)"] else 2
        for col_idx, col in enumerate(sheet.columns, 1):
            col_letter = get_column_letter(col_idx)
            header = str(sheet.cell(header_row, col_idx).value or "")
            if header in ["타겟 절대 경로 (URL)", "고위험 경로 (Endpoint)", "타겟 URL"]: sheet.column_dimensions[col_letter].width = 80  
            elif header == "발견된 JS 파일명": sheet.column_dimensions[col_letter].width = 50  
            elif header in ["탐지 사유", "Gemini AI 지능형 정보 노출 분석 가이드", "서브도메인 (Subdomain)"]: sheet.column_dimensions[col_letter].width = 55  
            elif header in ["📊 누적 / 🔥 신규 URL", "jsluice (누적 / 신규)", "🌟 서브도메인 (누적/신규)", "최초 발견일 (Wayback Machine)"]: sheet.column_dimensions[col_letter].width = 28
            elif header in ["응답 상태", "🔥 신규여부", "🌟 신규 서브", "🔮 잠재적 위험 확률", "📡 응답 상태"]: sheet.column_dimensions[col_letter].width = 18
            else: sheet.column_dimensions[col_letter].width = 18

    ws_dash.column_dimensions['B'].width = 35
    if "🔮 Gemini AI Ranking" in wb.sheetnames: wb["🔮 Gemini AI Ranking"].column_dimensions['D'].width = 80
    
    wb.save(f'reports/passive_recon_report_{now_str}.xlsx')
    
    with open(f'reports/postman_collection_{now_str}.json', 'w', encoding='utf-8') as f:
        json.dump(postman_collection, f, indent=4, ensure_ascii=False)
        
    print(f"[+] Postman API 및 엑셀 리포트 사출 완료!", flush=True)

if __name__ == '__main__':
    build_advanced_excel_report()
